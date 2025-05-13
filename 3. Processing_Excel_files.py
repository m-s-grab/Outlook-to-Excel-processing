import os #for creating file paths
import re #for regular expressions - cleaning NIP (Company ID)
import time #for measuring the length of the script
import shutil #for moving files between folders
import openpyxl #for editing Excel files
import pandas as pd #for using dataframes
import subprocess #for opening the master file at the end of the script
from datetime import datetime #for timestamps
import warnings #for supressing warnings from openpyxl (as the master file contains fields with data validation)
import json #for loading paths from config.json file
import sys #for checking if script is being used as exe

# Determines base path - config file should be in the same folder as exe/py scrypt
if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)  # Folder of the .exe
else:
    base_path = os.path.dirname(os.path.abspath(__file__))  # Folder of the .py

CONFIG_FILE = os.path.join(base_path, "config.json")

# Loads config - locations of the Data folder and supplier master file
with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
    config = json.load(f)

folder_base = config.get("processing_location")
supplier_file = config.get("supplier_file_location")

folder_to_process = os.path.join(folder_base, 'To_process')
folder_invalid = os.path.join(folder_base, 'Invalid_files')
folder_processed = os.path.join(folder_base, 'Processed')
folder_processed_msg = os.path.join(folder_processed, 'Processed_msg')
supplier_backup_file = os.path.join(folder_base, 'Supplier_backup.xlsm')
log_file = os.path.join(folder_base, 'log.txt')

start_col = 25
info_columns = [3, 5, 7]  # D, F, H

# Suppresses openpyxl warnings related to data validation issues
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
os.makedirs(folder_processed_msg, exist_ok=True)

# Ensures company ID (NIP) in the database is digits only, but many companies would provide
# their NIP "XXXXXX" in the form of "XXX XXX" or "ZZXXXXXX", etc.
def clean_nip(nip):
    return re.sub(r'\D', '', str(nip))

# Possible errors:
# 1. missing_x - supplier sending an Excel file with no services chosen
# 2. empty_category - supplier putting an "x" next to an empty ("-") category
def log_error(msg):
    with open(log_file, "a", encoding="utf-8") as log:
        log.write(msg + "\n")
    print(msg)

def check_errors(df):
    missing_x = True
    empty_category = False
    for _, row in df.iterrows():
        for i_service, i_x in [(2, 3), (4, 5), (6, 7)]:
            if str(row.iloc[i_x]).strip().lower() == 'x':
                missing_x = False
                if str(row.iloc[i_service]).strip() == '-':
                    empty_category = True
    return missing_x, empty_category

# As the stakeholder requires both msg file and its attachments to be saved,
# we ensure that msg files are always following xlsx files with the same name
def move_msg(xlsx_name, target_folder):
    msg_name = os.path.splitext(xlsx_name)[0] + ".msg"
    msg_path = os.path.join(folder_to_process, msg_name)
    if os.path.exists(msg_path):
        try:
            shutil.move(msg_path, os.path.join(target_folder, msg_name))
            print(f"📬 Moved MSG file: {msg_name}")
        except Exception as e:
            log_error(f"❗ Error moving MSG file {msg_name}: {e}")

# Starts measuring the time
start_time = time.time()
start_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
log_error(f"{start_timestamp}▶️ Started processing files from 'To_process' folder")

# Checks if master supplier file is open by trying to rename it - if it's open - end the script
try:
    os.rename(supplier_file, supplier_file)
except OSError:
    log_error("❌ Please close the master supplier file.")
    input("\nPress any key to exit...")
    sys.exit(1)

# Loads supplier workbook without removing existing macros
try:
    wb_supplier = openpyxl.load_workbook(supplier_file, keep_vba=True)
except Exception as e:
    log_error(f"❌ Unexpected error opening the supplier file: {e}")
    input("\nPress Enter to exit...")
    sys.exit(1)

sheet_data = wb_supplier['data1']
sheet_suppliers = wb_supplier['data2']

# Creates map of NIPs to rows
def create_nip_map(sheet, nip_column, start_row=4):
    mapping = {}
    for row in range(start_row, sheet.max_row + 1):
        sheet[f'{nip_column}{row}'].number_format = '@'
        val = sheet[f"{nip_column}{row}"].value
        if val:
            nip_clean = clean_nip(val)
            if nip_clean:
                mapping[nip_clean] = row
    return mapping

# Sets the column with NIPs (in this case - primary keys) and starting rows in each sheet
nip_map_data = create_nip_map(sheet_data, nip_column='F', start_row=3)
nip_map_suppliers = create_nip_map(sheet_suppliers, nip_column='H', start_row=3)

# Creates a list of invalid entries
invalid_entries = []

# Every xlsx file in the "To_process" folder is set to be processed
for file in os.listdir(folder_to_process):
    if not file.endswith('.xlsx'):
        continue

    path = os.path.join(folder_to_process, file)
    supplier_name = os.path.splitext(os.path.basename(file))[0].strip()

# Script opens each Excel file in "To_process" folder and checks for clean NIP in C7 cell.
    try:
        wb_info = openpyxl.load_workbook(path, data_only=True)
        wb_info['DATA']['C7'].number_format = '@'
        nip_raw = wb_info['DATA']['C7'].value
        nip_clean = clean_nip(nip_raw)
    except:
        log_error(f"❌ {file} - missing 'DATA' sheet")
        invalid_entries.append([supplier_name, 'Missing DATA', '', "Excel update", start_timestamp])
        shutil.move(path, os.path.join(folder_invalid, file))
        move_msg(file, folder_invalid)
        continue

# In case C7 did not contain any digits files are moved to "Invalid_files" folder and the "Invalid_files.xlsx" entry is updated
    if not nip_clean:
        log_error(f"⚠️ {file} - invalid NIP")
        invalid_entries.append([supplier_name, 'Invalid NIP', '', "Excel update", start_timestamp])
        shutil.move(path, os.path.join(folder_invalid, file))
        move_msg(file, folder_invalid)
        continue

# When NIP is found, the script:
# 1.Checks the "offer" sheet for errors
    try:
        df_offer = pd.read_excel(path, sheet_name='offer', header=None)
    except:
        log_error(f"❌ {file} - missing 'offer' sheet")
        invalid_entries.append([supplier_name, 'Missing offer sheet', nip_clean, "Excel update", start_timestamp])
        shutil.move(path, os.path.join(folder_invalid, file))
        move_msg(file, folder_invalid)
        continue

    missing_x, empty_category = check_errors(df_offer)
    if missing_x or empty_category:
        reason = 'no x' if missing_x else 'empty category'
        log_error(f"❌ {file} - error: {reason}")
        invalid_entries.append([supplier_name, reason, nip_clean, "Excel update", start_timestamp])
        shutil.move(path, os.path.join(folder_invalid, file))
        move_msg(file, folder_invalid)
        continue

    row_data = nip_map_data.get(nip_clean)
    row_suppliers = nip_map_suppliers.get(nip_clean)

    if not row_data or not row_suppliers:
        log_error(f"❌ {file} - error: Missing/invalid NIP")
        invalid_entries.append([supplier_name, 'Missing/invalid NIP', nip_clean, "Excel update", start_timestamp])
        shutil.move(path, os.path.join(folder_invalid, file))
        move_msg(file, folder_invalid)
        continue

    # 2. Checks if the row was already filled (based on entry in column 805 not being empty)
    # Column 805 contains the name of the xlsx file from which data was extracted in the particular row.
    # If the row was already filled, the script deletes all data from "X" cells before filling the row again.
    if sheet_data.cell(row=row_data, column=805).value:
        prev_file = sheet_data.cell(row=row_data, column=805).value
        log_error(f"🔁 Row {row_data} was overwritten (previously: {prev_file})")
        for col in range(25, 805):
            sheet_data.cell(row=row_data, column=col).value = None

    # Copies "x" values to master file one column after another (each column fills every 3rd cell)
    for i, column in enumerate(info_columns):
        for idx in range(min(190, len(df_offer) - 2)):
            try:
                val = str(df_offer.iloc[idx + 2, column]).strip().lower()
                if val == 'x':
                    col_index = start_col + i + idx * 3
                    sheet_data.cell(row=row_data, column=col_index).value = 'x'
            except Exception as e:
                log_error(f"❗ Błąd w pliku {file}, kol {column + 1}, row {idx + 4}: {e}")

    # Fills in the "Suppliers DATA" sheet in the master file (from "DATA" sheet in processed files)
    # Columns A, H, J, K, L contain some prefilled data (including NIP). They are skipped.
    mapping = {
        **{f'C{i}': col for i, col in zip(range(1, 7), 'BCDEFG')},
        **{f'C{i}': col for i, col in zip(range(8, 10), 'IM')},
        **{f'C{i}': col for i, col in zip(range(10, 22), 'NOPQRSTUVWXYZ')}
    }

    for source_cell, dest_col in mapping.items():
        val = wb_info['DATA'][source_cell].value
        sheet_suppliers[f"{dest_col}{row_suppliers}"] = val

    # Saves the name of the processed file in column 805
    sheet_data.cell(row=row_data, column=805).value = supplier_name

    # Moves the processed xlsx file (and the msg file) to "Processed" folder.
    shutil.move(path, os.path.join(folder_processed, file))
    move_msg(file, folder_processed_msg)
    log_error(f"✅ {file} successfully processed and saved to Excel.")

wb_supplier.save(supplier_file) # Saves the master file

# A path to an xlsx file saving info about all the invalid files
invalid_xlsx_path = os.path.join(folder_invalid, "0.Invalid.xlsx")

# If there are any new entries about invalid files, they are being added to new rows in "0.Invalid.xlsx"
# Info includes: Name of the file; Type of error; NIP; Script used (in case of this script it's "Excel update"); Timestamp.
if invalid_entries:
    try:
        df_new = pd.DataFrame(invalid_entries, columns=['File name', 'Message', 'NIP', 'Operation', 'Timestamp'])

        if os.path.exists(invalid_xlsx_path):
            df_existing = pd.read_excel(invalid_xlsx_path)
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            df_combined = df_new # If there is no "0.Invalid.xlsx" yet, it creates one

        df_combined['Timestamp'] = pd.to_datetime(df_combined['Timestamp'], errors='coerce')
        df_combined = df_combined.sort_values('Timestamp', ascending=False)
        df_unique = df_combined.drop_duplicates(subset='File name', keep='first') # Previous duplicates are removed

        df_unique.to_excel(invalid_xlsx_path, index=False)
        log_error(f"✅ Saved data to {invalid_xlsx_path} and removed duplicates.")
    except Exception as e:
        log_error(f"❌ Error saving to Excel: {e}")
else:
    log_error("ℹ️ No invalid entries to save to 0.Invalid.xlsx.")

# Opens the master file at the end of the script
subprocess.Popen([supplier_file], shell=True)

# Informs about completion of the processing and time taken
log_error(f"✅ Processing complete. File updated: {supplier_file}")
log_error(f"⏱️ Time taken: {time.time() - start_time:.2f} sec\n")

input("📄 Press any key to exit...")
