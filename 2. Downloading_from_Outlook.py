import os #for creating file paths
import win32com.client #for connecting to Outlook
from datetime import datetime, timedelta #for choosing a time period for emails
import openpyxl
from openpyxl import Workbook, load_workbook #for reading xlsx files without opening Excel
import shutil #for moving files
import warnings #for ignoring openpyxl warnings
import re #for regular expressions - cleaning NIP (Company ID)
import pandas as pd #for using dataframes
import json  #for reading json files
import sys #for checking if skript is activated as exe

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Determines base path - config file should be in the same folder as exe/py script
if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(base_path, "config.json")

# Loads config - location of the Data folder
with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
    config = json.load(f)

base_folder = config.get("processing_location")
to_process = os.path.join(base_folder, 'To_process')
invalid_folder = os.path.join(base_folder, 'Invalid_files')
tmp_folder = os.path.join(base_folder, "tmp")
invalid_xlsx_path = os.path.join(invalid_folder, "0.Invalid.xlsx")
log_path = os.path.join(base_folder, "mail_log.xlsx")

os.makedirs(to_process, exist_ok=True)
os.makedirs(invalid_folder, exist_ok=True)
os.makedirs(tmp_folder, exist_ok=True)

# Script asks how many hours back would the user like to check the mails
while True:
    try:
        hours = int(input("Enter the number of hours to check emails from: "))
        break
    except ValueError:
        print("Please enter a valid number of hours!")

time_limit = datetime.now() - timedelta(hours=hours)
execution_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# The user can choose to skipped messages that were already checked in the past
# The "mail_log" saves the data on messages already checked by the script
# (email of the sender; subject; date and time)

skip_logged = input("Skip emails already processed? (y/n): ").strip().lower() == 'y'

def read_log():
    if os.path.exists(log_path):
        return pd.read_excel(log_path)
    else:
        return pd.DataFrame(columns=["Email", "Subject", "Received"])

def write_log(email, subject, received):
    df = read_log()
    new_row = {"Email": email, "Subject": subject, "Received": received}
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_excel(log_path, index=False)

def is_logged(email, subject, received):
    df = read_log()
    return ((df["Email"] == email) & (df["Subject"] == subject) & (df["Received"] == received)).any()

def unique_filename(folder, filename):
    base, ext = os.path.splitext(filename)
    counter = 1
    unique = filename
    while os.path.exists(os.path.join(folder, unique)):
        unique = f"{base}_{counter}{ext}"
        counter += 1
    return unique

def log_invalid(filepath, row_data):
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.append(['File name', 'Message', 'NIP', 'Source', 'Execution time'])
    else:
        wb = load_workbook(filepath)
        ws = wb.active
    ws.append(row_data)
    wb.save(filepath)


# Connects to Outlook and checks on all the messages in Inbox.

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items
messages.Sort("[ReceivedTime]", True)

# Sets the counter for messages
total_msgs = 0
invalid_msgs = 0

# Main loop:
# 1. Checks if the message came within the chosen time period
for msg in messages:
    try:
        msg_time = msg.ReceivedTime.replace(tzinfo=None)
    except Exception:
        continue

    if msg_time < time_limit:
        break

    email = msg.SenderEmailAddress
    subject = msg.Subject
    time_str = msg_time.strftime("%Y-%m-%d %H:%M:%S")

    # 2. Checks if the message is already in the mail_log (if user chose it to)
    if skip_logged and is_logged(email, subject, time_str):
        continue

    # 3. Checks if the mail is a message from supplier sending back the filled Excel form.
    # If the message contains Excel attachments, it's temporarily saved in temp folder
    # If saved Excel contains two sheets - "DATA" and "offer" - script assumes it's a message from Supplier.
    attachments = msg.Attachments
    valid_attachments = []

    for att in attachments:
        filename = att.FileName
        if filename.endswith(('.xlsx', '.xlsm')):
            temp_path = os.path.join(tmp_folder, "tmp_" + filename)
            try:
                att.SaveAsFile(temp_path)
                wb = openpyxl.load_workbook(temp_path, data_only=True)
                if 'Supplier DATA' in wb.sheetnames and 'categories' in wb.sheetnames:
                    valid_attachments.append(temp_path)
                else:
                    wb.close()
                    os.remove(temp_path)
            except Exception:
                if os.path.exists(temp_path):
                    os.remove(temp_path)

    if not valid_attachments:
        write_log(email, subject, time_str)
        continue

    total_msgs += 1

    # 4. If the message came with more attachments that meet the criteria - it's moved to "Invalid_files" folder
    # If that's the case, all attachments are named after the name of the sender.
    if len(valid_attachments) > 1:
        try:
            sender = msg.SenderName.strip() if msg.SenderName else "unknown"
            base_name = f"{sender}_multiple"

            # 4a. Invalid attachments are checked for NIP (Company ID) in cell C7
            for file in valid_attachments:
                try:
                    wb = openpyxl.load_workbook(file, data_only=True)
                    nip = re.sub(r'\D', '', str(wb['Supplier DATA']['C7'].value or ""))
                    wb.close()

                    new_name = unique_filename(invalid_folder, f"{base_name}{os.path.splitext(file)[1]}")
                    shutil.move(file, os.path.join(invalid_folder, new_name))
                    print(f"⚠️ Moved to Invalid_files: {new_name}")

                    # 4b. Invalid messages' data is saved in the "0.Invalid.xlsx" log shared between the scripts.
                    log_invalid(invalid_xlsx_path, [os.path.splitext(new_name)[0], "More than one attachment", nip, "Outlook", execution_time])
                except Exception as e:
                    print(f"❌ Error with attachment: {e}")
                    if os.path.exists(file):
                        os.remove(file)

            #4c. The msg file is saved together with the Excel attachment.
            msg_name = unique_filename(invalid_folder, f"{base_name}.msg")
            msg.SaveAs(os.path.join(invalid_folder, msg_name))
            print(f"📬 Saved message as: {msg_name}")
            invalid_msgs += 1

        except Exception as e:
            print(f"❌ General error handling multiple attachments: {e}")

    #5. Valid messages are saved with a name of the Company from cell C1 in Supplier DATA sheet in the Excel attachment.
    else:
        valid = valid_attachments[0]
        try:
            wb = openpyxl.load_workbook(valid, data_only=True)
            name = str(wb['Supplier DATA']['C1'].value or "no_name").strip()
            wb.close()

            ext = os.path.splitext(valid)[1]
            msg_date = msg_time.strftime("%d-%m-%Y")
            final_name = unique_filename(to_process, f"{name}_cat_{msg_date}{ext}")

            # 5a. The attachment is moved to the "To_process" folder for another script to work with it.
            shutil.move(valid, os.path.join(to_process, final_name))
            print(f"✅ Moved to To_process: {final_name}")

            # 5b. The msg file is saved together with the attachment.
            msg_name = unique_filename(to_process, os.path.splitext(final_name)[0] + ".msg")
            msg.SaveAs(os.path.join(to_process, msg_name))
            print(f"📬 Saved message as: {msg_name}")
        except Exception as e:
            print(f"❌ Error processing valid file: {e}")
            if os.path.exists(valid):
                os.remove(valid)

    write_log(email, subject, time_str)

# ✅ Summary
print(f"\n🔚 Emails matching criteria: {total_msgs}\n❌ Invalid messages: {invalid_msgs}")
